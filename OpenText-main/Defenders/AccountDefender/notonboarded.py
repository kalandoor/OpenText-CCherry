import requests
import json
import openpyxl
import os
import csv
import configparser
from openpyxl.styles import PatternFill

def get_auth_token():
    # Function to get authentication token
    # Add error handling
    try:
        config = configparser.ConfigParser()
        config.read('config.ini')
        base_url = config.get('prismacloud', 'api_url')
        username = config.get('prismacloud', 'username')
        password = config.get('prismacloud', 'password')
        
        url = f"{base_url}/login"
        payload = {
          "username": username,
          "password": password
        }
        headers = {
          'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=json.dumps(payload))
        response_json = response.json()
        if 'token' in response_json:
            return response_json['token']
        else:
            print("Failed to get authentication token")
            return None
    except Exception as e:
        print(f"Failed to get authentication token: {e}")
        return None
    
token = get_auth_token()

def get_account_list():
    # Function to retrieve account list
    try:
        config = configparser.ConfigParser()
        config.read('config.ini')
        base_url = config.get('prismacloud', 'api_url')
        
        url = f"{base_url}/cloud/group"
        headers = {
            'Accept': 'application/json; charset=UTF-8',
            'x-redlock-auth': token
        }
        info = requests.get(url, headers=headers)
        if info.status_code == 200:
            with open('info.json', 'w') as file:
                json.dump(info.json(), file)
        else:
            print('Failed to retrieve cloud accounts:', info.status_code, info.text)
    except Exception as e:
        print(f"Failed to retrieve cloud accounts: {e}")

def get_defender_list():
    # Function to retrieve defender list
    try:
        config = configparser.ConfigParser()
        config.read('config.ini')
        console_url = config.get('prismacloud', 'console_url')
        
        url = f"{console_url}/api/v32.04/defenders/download"
        headers = {
            'x-redlock-auth': token
        }
        defender = requests.get(url, headers=headers)
        if defender.status_code == 200:
            csv_data = defender.text.splitlines()
            csv_reader = csv.reader(csv_data)
            with open('defender_list.csv', mode='w', newline='') as file:
                writer = csv.writer(file)
                for row in csv_reader:
                    writer.writerow(row)
            print("CSV file saved successfully.")
        else:
            print("Failed to fetch defender list.")
    except Exception as e:
        print(f"Failed to fetch defender list: {e}")

def find_defender_accounts():
    # Function to find defender accounts
    try:
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        headers = {
            'aws': ['Account ID', 'Onboarded'],
            'azure': ['Account ID', 'Onboarded'],
            'gcp': ['Account ID', 'Onboarded']
        }

        for sheet_name, sheet_headers in headers.items():
            sheet = wb.create_sheet(title=sheet_name)
            sheet.append(sheet_headers)

        with open('defender_list.csv', 'r') as csvfile:
            csvreader = csv.DictReader(csvfile)
            for row in csvreader:
                account_id = row['Account ID']
                hostname = row['Hostname']

                if hostname.startswith("gke"):
                    sheet = wb['gcp']
                elif hostname.startswith("aks"):
                    sheet = wb['azure']
                elif hostname.startswith("ip"):
                    sheet = wb['aws']
                else:
                    continue

                account_ids_sheet = {sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)}
                if account_id not in account_ids_sheet:
                    sheet.append([account_id])

            for sheet_name in ['aws', 'azure', 'gcp']:
                sheet = wb[sheet_name]
                for row in range(2, sheet.max_row + 1):
                    if not sheet.cell(row=row, column=1).value:
                        sheet.delete_rows(row, 1)

        wb.save('notonboarded.xlsx')
    except Exception as e:
        print(f"Failed to find defender accounts: {e}")

def confirm_onboarded():
    # Function to confirm onboarded accounts
    try:
        with open('info.json', 'r') as file:
            info_data = json.load(file)

        wb = openpyxl.load_workbook('notonboarded.xlsx')

        for sheet_name in ['aws', 'azure', 'gcp']:
            sheet = wb[sheet_name]
            for row in range(2, sheet.max_row + 1):
                account_id = sheet.cell(row=row, column=1).value
                sheet.cell(row=row, column=2).value = 'No'
                for entry in info_data:
                    for account in entry.get('accounts', []):
                        if account_id == account.get('id') or account_id in account.get('name', ''):
                            sheet.cell(row=row, column=2).value = 'Yes'
                            break
                    if sheet.cell(row=row, column=2).value == 'Yes':
                        break
                else:
                    sheet.cell(row=row, column=2).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        wb.save('notonboarded.xlsx')
    except Exception as e:
        print(f"Failed to confirm onboarded accounts: {e}")

def add_total():
    # Function to add total not-onboarded count
    try:
        wb = openpyxl.load_workbook('notonboarded.xlsx')
        for sheet_name in ['aws', 'azure', 'gcp']:
            sheet = wb[sheet_name]
            header = 'Total Not-Onboarded'
            sheet['D1'] = header
            no_count = sum(1 for cell in sheet['B'] if cell.value == 'No')
            sheet['D2'] = no_count
        wb.save('notonboarded.xlsx')
    except Exception as e:
        print(f"Failed to add total not")

get_account_list()
get_defender_list()
find_defender_accounts()
confirm_onboarded()
add_total()

#Remove Files
os.remove("./info.json")
os.remove("./defender_list.csv")