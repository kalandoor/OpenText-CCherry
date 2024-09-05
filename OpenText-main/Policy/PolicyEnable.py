import requests
import json
import openpyxl
import os
import configparser

config = configparser.ConfigParser()
config.read('config.ini')
base_url = config.get('prismacloud', 'api_url')
username = config.get('prismacloud', 'username')
password = config.get('prismacloud', 'password')

# Function to get authentication token
def get_auth_token():
    url = f"{base_url}/login"
    payload = {
      "username": username,
      "password": password
    }
    headers = {
      'Content-Type': 'application/json'
    }

    response = requests.request("POST", url, headers=headers, data=json.dumps(payload))
    return response.json().get("token", "")

# Save the token in the 'token' variable
token = get_auth_token()

# Pull Policies for the Compliance Standard and Save it as a .json file
def get_standards(compliance_standard, file_name):
    url = f"{base_url}/v2/policy?policy.complianceStandard={compliance_standard}"
    payload={}
    headers = {
    'Accept': 'application/json; charset=UTF-8',
    'x-redlock-auth': token
    }

    response = requests.request("GET", url, headers=headers, data=payload)

    # Check if the response is successful (status code 200) and save it to a JSON file
    if response.status_code == 200:
        with open(file_name, 'w') as file:
            file.write(response.text)
        print(f"Response saved to '{file_name}'")
    else:
        print(f"Failed to get {compliance_standard} standards. Status code: {response.status_code}")

# Get standards for AWS
get_standards("Opentext Standard for AWS", 'aws.json')

# # Get standards for Azure
get_standards("Opentext Standard for Azure", 'azure.json')

# # Get standards for GCP
get_standards("Opentext Standard for GCP", 'gcp.json')

# Create Excel sheets for each JSON file
def create_excel_sheets():
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)  # Remove default sheet

    for file_name in ['aws.json', 'azure.json', 'gcp.json']:
        if os.path.exists(file_name):
            with open(file_name, 'r') as file:
                data = json.load(file)
                sheet_name = file_name.split('.')[0]
                ws = wb.create_sheet(title=sheet_name)

                # Write header
                header = [
                    "policyId",
                    "name",
                    "policyType",
                    "policySubTypes",
                    "description",
                    "severity",
                    "cloudType",
                    "complianceMetadata",
                    "enabled"
                ]
                ws.append(header)

                # Write data
                for item in data:
                    row_data = [
                        item.get('policyId', ''),
                        item.get('name', ''),
                        item.get('policyType', ''),
                        ', '.join(item.get('policySubTypes', [])),
                        item.get('description', ''),
                        item.get('severity', ''),
                        item.get('cloudType', ''),
                        json.dumps(item.get('complianceMetadata', {})),
                        item.get('enabled', '')
                    ]
                    ws.append(row_data)

        else:
            print(f"File {file_name} does not exist.")

    wb.save('policies.xlsx')
    print("Excel file created: policies.xlsx")

create_excel_sheets()

# Remove .json files as they are no longer needed.
def remove_json_files():
    for file_name in ['aws.json', 'azure.json', 'gcp.json']:
        if os.path.exists(file_name):
            os.remove(file_name)
            print(f"File {file_name} removed.")
        else:
            print(f"File {file_name} does not exist.")

remove_json_files()

# This is to Enable the policies that are labled false in the Policies.xlsx file
def enable_policy():
    wb = openpyxl.load_workbook('policies.xlsx')
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=2, max_col=9, values_only=True):
            policyid, _, _, _, _, _, _, _, enabled = row
            if not enabled:  # Column I is false
                url = f"{base_url}/policy/{policyid}"
                payload = "{\n  \"enabled\": true\n}"
                headers = {
                    'Content-Type': 'application/json; charset=UTF-8',
                    'Accept': 'application/json; charset=UTF-8',
                    'x-redlock-auth': token
                }
                response = requests.request("PUT", url, headers=headers, data=payload)
                if response.status_code == 200:
                    print(f"Policy {policyid} in sheet {sheet_name} updated successfully.")
                else:
                    print(f"Failed to update policy {policyid} in sheet {sheet_name}. Status code: {response.status_code}")

enable_policy()   