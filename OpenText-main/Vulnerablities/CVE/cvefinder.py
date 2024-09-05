import requests
import json
import os
import openpyxl
import configparser

config = configparser.ConfigParser()
config.read('config.ini')
base_url = config.get('prismacloud', 'api_url')
username = config.get('prismacloud', 'username')
password = config.get('prismacloud', 'password')
console_url = config.get('prismacloud', 'console_url')

# Function to get authentication token
def get_auth_token():
    url = f"{base_url}/login"
    payload = {
      "username": username,
      "password": password
    }
    headers = {
      'Content-Type': 'application/json'
    }

    response = requests.request("POST", url, headers=headers, data=json.dumps(payload))
    response_json = response.json()
    if 'token' in response_json:
        return response_json['token']
    else:
        print("Failed to get authentication token")
        return None

# Save the token in the 'token' variable
token = get_auth_token()

def get_cve_stats():
    if token is None:
        print("Authentication token not available. Exiting.")
        return

    cve_ID = input("Which CVE? ")

    url = f"{console_url}/api/v1/stats/vulnerabilities/impacted-resources?cve={cve_ID}"
    payload={}
    headers = {
        'Accept': 'application/json',
        'x-redlock-auth': token
    }

    response = requests.request("GET", url, headers=headers, data=payload)
    if response.status_code == 200:
        with open('response.json', 'w') as file:
            json.dump(response.json(), file)

        # Load response data
        response_data = response.json()

        # Create a new Excel workbook
        workbook = openpyxl.Workbook()

        # Remove default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        # Create a new sheet called "total"
        total_sheet = workbook.create_sheet("total")

        # Define headers
        headers = ["imagesCount", "hostsCount", "functionsCount", "registryImagesCount"]

        # Write headers to "total" sheet
        total_sheet.append(headers)

        # Extract and write values to "total" sheet
        values = [
            response_data["imagesCount"],
            response_data["hostsCount"],
            response_data["functionsCount"],
            response_data["registryImagesCount"]
        ]
        total_sheet.append(values)

        # Save the workbook
        workbook.save('CVEReview.xlsx')

    else:
        print(f"Failed to fetch CVE stats. Status code: {response.status_code}")

get_cve_stats()

def load_response_json():
    with open('response.json', 'r') as file:
        return json.load(file)

def get_image_ids():
    response_data = load_response_json()
    image_ids = set()
    for image in response_data.get("images", []):
        for container in image.get("containers", []):
            image_id = container.get("imageID", "")
            # Add only unique image IDs to the set
            image_ids.add(image_id)
    return list(image_ids)
    
image_ids = get_image_ids()
# Split image_ids into chunks of 100
chunk_size = 100
chunks = [image_ids[i:i+chunk_size] for i in range(0, len(image_ids), chunk_size)]

# Iterate over each chunk
for i, chunk in enumerate(chunks, start=1):
    # Convert chunk to comma-separated string
    id_str = ','.join(chunk)
    url = f"{console_url}/api/v1/images?id={id_str}"

    payload = {}
    headers = {
      'Accept': 'application/json',
      'x-redlock-auth': token
    }

    try:
        image_data = requests.request("GET", url, headers=headers, data=payload)
        image_data.raise_for_status()  # Raise an exception for non-200 status codes
        with open(f'image_data{i}.json', 'w') as file:
            json.dump(image_data.json(), file)
    except Exception as e:
        print(f"An error occurred: {e}")

def load_image_data():
    image_data_list = []
    i = 1
    while True:
        try:
            with open(f'image_data{i}.json', 'r') as file:
                image_data_list.extend(json.load(file))
            i += 1
        except FileNotFoundError:
            break
    return image_data_list

def load_excel_workbook():
    return openpyxl.load_workbook('CVEReview.xlsx')

def put_data_into_excel():
    image_data_list = load_image_data()
    workbook = load_excel_workbook()

    # Create new sheets
    gcp_sheet = workbook.create_sheet("GCP")
    azure_sheet = workbook.create_sheet("Azure")
    aws_sheet = workbook.create_sheet("AWS")

    # Write headers
    headers = ["Image", "Provider", "AccountID", "Created", "Namespaces", "Clusters", "Instances", "Hosts", "Host Name"]
    gcp_sheet.append(headers)
    azure_sheet.append(headers)
    aws_sheet.append(headers)

    # Map provider names to their respective sheets
    provider_sheet_map = {
        "gcp": gcp_sheet,
        "azure": azure_sheet,
        "aws": aws_sheet
    }

    for image_data in image_data_list:
        cloud_metadata = image_data.get("cloudMetadata", {})
        provider = cloud_metadata.get("provider", "")
        account_id = cloud_metadata.get("accountID", "")
        created = image_data.get("image", {}).get("created", "")
        namespaces = image_data.get("namespaces", [""])
        clusters = image_data.get("clusters", [""])
        instances_count = len(image_data.get("instances", []))
        host_name = set(instance.get("host", "") for instance in image_data.get("instances", []))

        # Get the sheet based on the provider, or skip if unknown
        sheet = provider_sheet_map.get(provider)
        if sheet is None:
            continue

        # Write data to the sheet
        sheet.append([
            cloud_metadata.get("image", ""),  # Use get() with a default value
            provider,
            account_id,
            created,
            ",".join(namespaces),
            ",".join(clusters),
            instances_count,
            len(image_data.get("hosts", [])),
            str(host_name)
        ])

    workbook.save('CVEReview.xlsx')

# Call the function to put data into the Excel sheet
put_data_into_excel()

# Remove the files
# Assuming the files are in the current directory
file_directory = "./"
for i in range(1, 100):  # Assuming a maximum of 100 files
    file_path = os.path.abspath(f"{file_directory}image_data{i}.json")
    if os.path.exists(file_path):
        os.remove(file_path)
    else:
        break  # Exit the loop if a file is not found
os.remove("./response.json")
